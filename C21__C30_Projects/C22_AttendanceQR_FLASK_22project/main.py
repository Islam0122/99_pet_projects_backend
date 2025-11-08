from flask import Flask, request, render_template_string, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import requests
import logging

app = Flask(__name__)
FILE_NAME = "attendance.xlsx"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active

        headers = ["Имя", "Действие", "Дата и время", "IP", "Координаты", "Адрес"]
        ws.append(headers)

        header_fill = PatternFill(start_color="115e59", end_color="115e59", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 50

        wb.save(FILE_NAME)
        logger.info("Excel файл создан и отформатирован")


def get_address(lat, lon):
    try:
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {"lat": lat, "lon": lon, "format": "json", "accept-language": "ru"}
        headers = {"User-Agent": "IslamDevAttendanceSystem/1.0"}
        response = requests.get(url, params=params, headers=headers, timeout=5)

        if response.status_code == 200:
            data = response.json()
            address = data.get("display_name", "Не определено")
            logger.info(f"Адрес определён: {address[:50]}...")
            return address
        else:
            logger.warning(f"Ошибка геокодирования: {response.status_code}")
    except requests.exceptions.Timeout:
        logger.error("Timeout при запросе геокодирования")
    except Exception as e:
        logger.error(f"Ошибка при получении адреса: {e}")

    return "Не определено"


@app.route("/")
def index():
    try:
        with open("index.html", encoding="utf-8") as f:
            html = f.read()
        return render_template_string(html)
    except FileNotFoundError:
        return jsonify({"error": "index.html не найден"}), 404


@app.route("/mark", methods=["POST"])
def mark():
    try:
        data = request.get_json()

        if not data:
            return jsonify({"error": "Нет данных"}), 400

        name = data.get("name", "Неизвестный").strip()
        action_type = data.get("type")
        lat = data.get("lat")
        lon = data.get("lon")
        ip = request.remote_addr
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not name or name == "Неизвестный":
            return "Ошибка: имя не указано", 400

        if action_type not in ["in", "out"]:
            return "Ошибка: неверный тип действия", 400

        action_text = "Пришёл" if action_type == "in" else "Ушёл"

        address = get_address(lat, lon) if lat and lon else "Не определено"
        coords = f"{lat}, {lon}" if lat and lon else "Не определено"

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        row_data = [name, action_text, timestamp, ip, coords, address]
        ws.append(row_data)

        row_num = ws.max_row

        action_fill = PatternFill(
            start_color="d1fae5" if action_type == "in" else "fee2e2",
            end_color="d1fae5" if action_type == "in" else "fee2e2",
            fill_type="solid"
        )
        ws[f'B{row_num}'].fill = action_fill

        for cell in ws[row_num]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        wb.save(FILE_NAME)

        logger.info(f"{name} - {action_text} в {timestamp}")

        return f"{name}, вы {action_text.lower()} в {timestamp}"

    except Exception as e:
        logger.error(f"Ошибка при обработке запроса: {e}")
        return f"Ошибка сервера: {str(e)}", 500


@app.route("/stats")
def stats():
    try:
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        total = ws.max_row - 1
        return jsonify({
            "total_records": total,
            "file": FILE_NAME,
            "status": "ok"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Страница не найдена"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Внутренняя ошибка сервера"}), 500


if __name__ == "__main__":
    init_excel()
    logger.info("Запуск Islam Dev Attendance System...")
    logger.info(f"Excel файл: {FILE_NAME}")
    app.run(host="0.0.0.0", port=5001, debug=False)